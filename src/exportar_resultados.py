"""
Módulo encargado de exportar los resultados obtenidos
en un archivo .xlsx con dos hojas:
1. Flujos: DataFrame con las columnas originales y las columnas calculadas.
2. Resumen: VEP base, VEP de cada escenario de choque y variación del VEP respecto al escenario base.
"""

# Importaciones necesarias
import pandas as pd

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def exportar_resultados(df_flujos: pd.DataFrame, resultados: dict, ruta_salida: str) -> None:
    
    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:

        # Exportar DataFrame de flujos
        df_flujos.to_excel(writer, sheet_name="Flujos", index=False)

        # Crear tabla estructurada de análisis
        vep_base = resultados["VEP Base"]

        escenarios = ["Paralelo Arriba",
                        "Paralelo Abajo",
                        "Empinamiento",
                        "Aplanamiento",
                        "Corto Arriba",
                        "Corto Abajo"]

        filas_resumen = []

        # Caso base
        filas_resumen.append({"Escenario": "Base",
                                "VEP": vep_base,
                                "Delta VEP": 0,
                                "Delta VEP %": 0})

        for escenario in escenarios:
            vep = resultados[f"VEP {escenario}"]
            delta_vep = resultados[f"Delta VEP {escenario}"]

            filas_resumen.append({"Escenario": escenario,
                                    "VEP": vep,
                                    "Delta VEP": delta_vep,
                                    "Delta VEP %": delta_vep / vep_base})

        df_resumen = pd.DataFrame(filas_resumen)

        # Exportar resumen
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)

        # Obtener hojas de Excel
        hoja_flujos = writer.sheets["Flujos"]
        hoja_resumen = writer.sheets["Resumen"]

        # Estilo de encabezados
        relleno_gris = PatternFill(fill_type="solid", fgColor="808080")

        fuente_encabezado = Font(bold=True, color="FFFFFF")

        # Aplicar formato a ambas hojas
        for hoja in (hoja_flujos, hoja_resumen):

            # Formato de encabezados
            for celda in hoja[1]:
                celda.fill = relleno_gris
                celda.font = fuente_encabezado
                celda.alignment = Alignment(horizontal="center", vertical="center")

            # Ajustar ancho de columnas
            for columna in hoja.columns:
                longitud_maxima = 0
                letra_columna = get_column_letter(columna[0].column)

                for celda in columna:
                    if celda.value is not None:
                        longitud_maxima = max(longitud_maxima, len(str(celda.value)))

                hoja.column_dimensions[letra_columna].width = (longitud_maxima + 2)

        # Agregar filtros a la tabla de flujos
        hoja_flujos.auto_filter.ref = hoja_flujos.dimensions

        # Convertir porcentajes en la hoja de resumen
        for fila in range(2, hoja_resumen.max_row + 1):
            hoja_resumen.cell(row=fila, column=4).number_format = "0.00%"

        # Convertir valores monetarios en la hoja de resumen
        for fila in range(2, hoja_resumen.max_row + 1):
            hoja_resumen.cell(row=fila, column=2).number_format = '#,##0.00'
            hoja_resumen.cell(row=fila, column=3).number_format = '#,##0.00'